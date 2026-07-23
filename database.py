import os
import hashlib
from datetime import datetime

# Use PostgreSQL if DATABASE_URL is set (Render), otherwise SQLite for local dev
DATABASE_URL = os.environ.get("DATABASE_URL", "")

if DATABASE_URL:
    import psycopg2
    import psycopg2.extras

    PG_SSLMODE = os.environ.get("PGSSLMODE", "require")

    def get_connection():
        conn = psycopg2.connect(DATABASE_URL, sslmode=PG_SSLMODE)
        conn.autocommit = False
        return conn

    def _fetchone_dict(cur):
        if cur.description is None:
            return None
        cols = [d[0] for d in cur.description]
        row = cur.fetchone()
        if row is None:
            return None
        return dict(zip(cols, row))

    def _fetchall_dict(cur):
        if cur.description is None:
            return []
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]

else:
    import sqlite3

    DB_PATH = os.environ.get("DB_PATH", os.path.join(os.path.dirname(__file__), "tasks.db"))

    def get_connection():
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
        return conn


def init_db():
    conn = get_connection()
    if DATABASE_URL:
        cur = conn.cursor()
        cur.execute("""
        CREATE TABLE IF NOT EXISTS features (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            release_date TEXT DEFAULT '',
            comments TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        )""")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS tasks (
            id TEXT PRIMARY KEY,
            content TEXT NOT NULL,
            type TEXT DEFAULT '',
            date TEXT DEFAULT '',
            owner TEXT DEFAULT '',
            status TEXT DEFAULT 'To Do',
            approval TEXT DEFAULT 'Draft',
            comment TEXT DEFAULT '',
            link TEXT DEFAULT '',
            priority TEXT DEFAULT '',
            feature_id TEXT DEFAULT '',
            source TEXT DEFAULT '',
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )""")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS task_platforms (
            task_id TEXT NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
            platform TEXT NOT NULL,
            PRIMARY KEY (task_id, platform)
        )""")
        cur.execute("CREATE TABLE IF NOT EXISTS platforms (name TEXT PRIMARY KEY)")
        cur.execute("CREATE TABLE IF NOT EXISTS owners (name TEXT PRIMARY KEY)")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS import_log (
            id SERIAL PRIMARY KEY,
            filename TEXT,
            imported_at TEXT,
            row_count INTEGER
        )""")
        conn.commit()
        cur.close()
        # Migrations for PostgreSQL (add columns if missing)
        cur = conn.cursor()
        try:
            cur.execute("ALTER TABLE features ADD COLUMN comments TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            conn.rollback()
        try:
            cur.execute("ALTER TABLE tasks ADD COLUMN feature_id TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            conn.rollback()
        cur.close()
    else:
        conn.executescript("""
        CREATE TABLE IF NOT EXISTS features (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            release_date TEXT DEFAULT '',
            comments TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS tasks (
            id TEXT PRIMARY KEY,
            content TEXT NOT NULL,
            type TEXT DEFAULT '',
            date TEXT DEFAULT '',
            owner TEXT DEFAULT '',
            status TEXT DEFAULT 'To Do',
            approval TEXT DEFAULT 'Draft',
            comment TEXT DEFAULT '',
            link TEXT DEFAULT '',
            priority TEXT DEFAULT '',
            feature_id TEXT DEFAULT '',
            source TEXT DEFAULT '',
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS task_platforms (
            task_id TEXT NOT NULL,
            platform TEXT NOT NULL,
            FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE CASCADE,
            PRIMARY KEY (task_id, platform)
        );
        CREATE TABLE IF NOT EXISTS platforms (name TEXT PRIMARY KEY);
        CREATE TABLE IF NOT EXISTS owners (name TEXT PRIMARY KEY);
        CREATE TABLE IF NOT EXISTS import_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT,
            imported_at TEXT,
            row_count INTEGER
        );
        """)
        conn.commit()
        # Add feature_id column if missing (migration for existing DBs)
        try:
            conn.execute("ALTER TABLE tasks ADD COLUMN feature_id TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            pass
        # Add comments column to features if missing
        try:
            conn.execute("ALTER TABLE features ADD COLUMN comments TEXT DEFAULT ''")
            conn.commit()
        except Exception:
            pass

    # Seed defaults
    default_platforms = ['Instagram', 'LinkedIn', 'YouTube', 'WhatsApp', 'Blog',
                         'Customer.io', 'Instantly', 'Twitter/X', 'Facebook', 'Loom']
    for p in default_platforms:
        _upsert_ignore(conn, "INSERT INTO platforms (name) VALUES (%s)", "INSERT OR IGNORE INTO platforms (name) VALUES (?)", (p,))

    default_owners = ['Mahek', 'Charu', 'Samriddhi', 'Aamir', 'Winston', 'Akshit']
    for o in default_owners:
        _upsert_ignore(conn, "INSERT INTO owners (name) VALUES (%s)", "INSERT OR IGNORE INTO owners (name) VALUES (?)", (o,))

    conn.commit()
    conn.close()


def _upsert_ignore(conn, pg_sql, sqlite_sql, params):
    if DATABASE_URL:
        cur = conn.cursor()
        try:
            cur.execute(pg_sql + " ON CONFLICT DO NOTHING", params)
        except Exception:
            conn.rollback()
        cur.close()
    else:
        conn.execute(sqlite_sql, params)


def _query_one(conn, pg_sql, sqlite_sql, params=()):
    if DATABASE_URL:
        cur = conn.cursor()
        cur.execute(pg_sql, params)
        result = _fetchone_dict(cur)
        cur.close()
        return result
    else:
        row = conn.execute(sqlite_sql, params).fetchone()
        return dict(row) if row else None


def _query_all(conn, pg_sql, sqlite_sql, params=()):
    if DATABASE_URL:
        cur = conn.cursor()
        cur.execute(pg_sql, params)
        result = _fetchall_dict(cur)
        cur.close()
        return result
    else:
        rows = conn.execute(sqlite_sql, params).fetchall()
        return [dict(r) for r in rows]


def _query_scalar(conn, sql_pg, sql_sqlite, params=()):
    if DATABASE_URL:
        cur = conn.cursor()
        cur.execute(sql_pg, params)
        row = cur.fetchone()
        cur.close()
        return row[0] if row else 0
    else:
        row = conn.execute(sql_sqlite, params).fetchone()
        return row[0] if row else 0


# --- Platform normalization ---
PLATFORM_MAP = {
    'instagram': 'Instagram', 'linkedin': 'LinkedIn', 'youtube': 'YouTube',
    'yt': 'YouTube', 'whatsapp': 'WhatsApp', 'whtsapp': 'WhatsApp',
    'wa': 'WhatsApp', 'blog': 'Blog', 'customer.io': 'Customer.io',
    'instantly': 'Instantly', 'twitter/x': 'Twitter/X', 'twitter': 'Twitter/X',
    'facebook': 'Facebook', 'loom': 'Loom', 'update': 'WhatsApp',
    'updates': 'WhatsApp', 'update grp': 'WhatsApp',
}


def normalize_platforms(raw_platform_str):
    if not raw_platform_str or raw_platform_str.strip() == '':
        return []
    raw = raw_platform_str.lower().strip()
    parts = []
    for sep in ['/', ',', '&']:
        if sep in raw:
            parts = [p.strip() for p in raw.split(sep) if p.strip()]
            break
    if not parts:
        parts = [raw]
    result = set()
    for part in parts:
        part = part.strip()
        for noise in ['update group', 'update grp', 'vid', 'video']:
            part = part.replace(noise, '').strip()
        if part in PLATFORM_MAP:
            result.add(PLATFORM_MAP[part])
        else:
            matched = False
            for key, val in PLATFORM_MAP.items():
                if key in part or part in key:
                    result.add(val)
                    matched = True
                    break
            if not matched and part:
                result.add(part.title())
    return sorted(result)


def normalize_status(raw):
    if not raw:
        return 'To Do'
    s = raw.strip()
    mapping = {'done': 'Done', 'in progress': 'In Progress', 'to do': 'To Do',
               'carryforward': 'Carryforward', 'carryfoward': 'Carryforward',
               'blocked': 'Blocked', 'completed': 'Done', '-': 'To Do'}
    return mapping.get(s.lower(), s)


def normalize_approval(raw):
    if not raw:
        return 'Draft'
    s = raw.strip()
    mapping = {'approved': 'Approved', 'to be approved': 'To be approved',
               'in progress': 'In Progress', 'rejected': 'Rejected', 'draft': 'Draft'}
    return mapping.get(s.lower(), s)


def normalize_type(raw):
    if not raw:
        return ''
    s = raw.strip()
    mapping = {'content': 'Content', 'campaign': 'Campaign', 'camoaign': 'Campaign',
               'carousel': 'Carousel', 'reel': 'REEL', 'teaser video': 'Teaser Video',
               'internal prep': 'Internal Prep', 'email campaign': 'Email Campaign',
               'kite': 'Content', 'linkedin': 'Content', 'customer.io': 'Campaign',
               'research': 'Research', 'creatives': 'Creatives', 'documentation': 'Documentation'}
    return mapping.get(s.lower(), s.title())


def make_task_id(content, date, owner):
    raw = f"{content}|{date}|{owner}"
    return "xl_" + hashlib.md5(raw.encode()).hexdigest()[:12]


def import_from_excel(filepath):
    import openpyxl
    import uuid as _uuid
    wb = openpyxl.load_workbook(filepath, data_only=True)
    conn = get_connection()
    now = datetime.now().isoformat()
    imported_count = 0

    if 'Marketing Calendar' in wb.sheetnames:
        ws = wb['Marketing Calendar']

        # First pass: collect features from column 8 and create them
        feature_names = set()
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if not row[0] or not isinstance(row[0], datetime):
                continue
            feat = str(row[8] or '').strip()
            if feat and feat != 'None':
                feature_names.add(feat)

        # Get existing features
        existing_feats = _query_all(conn, "SELECT * FROM features", "SELECT * FROM features")
        existing_map = {f['name'].lower(): f['id'] for f in existing_feats}
        feature_map = dict(existing_map)  # lowercase name -> id

        for feat_name in sorted(feature_names):
            if feat_name.lower() not in feature_map:
                feat_id = f"feat_{_uuid.uuid4().hex[:8]}"
                if DATABASE_URL:
                    cur = conn.cursor()
                    cur.execute("INSERT INTO features (id, name, release_date, comments, created_at) VALUES (%s, %s, %s, %s, %s)",
                                (feat_id, feat_name, '', '', now))
                    cur.close()
                else:
                    conn.execute("INSERT INTO features (id, name, release_date, comments, created_at) VALUES (?, ?, ?, ?, ?)",
                                 (feat_id, feat_name, '', '', now))
                feature_map[feat_name.lower()] = feat_id

        conn.commit()

        # Second pass: import tasks with correct column mapping
        # [0]=Date [1]=Day [2]=Content [3]=Type [4]=Platform [5]=Platform2 [6]=Status [7]=Owner [8]=Feature [9]=FeatureType [10]=Approval [11]=Comment [12]=Link
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if not row[0] or not isinstance(row[0], datetime):
                continue
            date_str = row[0].strftime('%Y-%m-%d')
            content = str(row[2] or '').strip()
            if not content:
                continue

            task_type = normalize_type(str(row[3] or ''))
            platforms = normalize_platforms(str(row[4] or ''))
            status = normalize_status(str(row[6] or ''))
            owner = str(row[7] or '').strip()
            feat_name = str(row[8] or '').strip()
            approval = normalize_approval(str(row[10] or ''))
            comment = str(row[11] or '').strip()
            link = str(row[12] or '').strip() if len(row) > 12 else ''
            if link == 'None':
                link = ''

            feature_id = feature_map.get(feat_name.lower(), '') if feat_name and feat_name != 'None' else ''

            task_id = make_task_id(content, date_str, owner)

            if DATABASE_URL:
                cur = conn.cursor()
                cur.execute("""
                    INSERT INTO tasks (id, content, type, date, owner, status, approval, comment, link, priority, feature_id, source, created_at, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (id) DO UPDATE SET
                        content=EXCLUDED.content, type=EXCLUDED.type, date=EXCLUDED.date,
                        owner=EXCLUDED.owner, status=EXCLUDED.status, approval=EXCLUDED.approval,
                        comment=EXCLUDED.comment, link=EXCLUDED.link, feature_id=EXCLUDED.feature_id, updated_at=EXCLUDED.updated_at
                """, (task_id, content, task_type, date_str, owner, status, approval, comment, link, '', feature_id, 'Marketing Calendar', now, now))
                cur.execute("DELETE FROM task_platforms WHERE task_id = %s", (task_id,))
                for p in platforms:
                    cur.execute("INSERT INTO task_platforms (task_id, platform) VALUES (%s, %s) ON CONFLICT DO NOTHING", (task_id, p))
                    cur.execute("INSERT INTO platforms (name) VALUES (%s) ON CONFLICT DO NOTHING", (p,))
                if owner:
                    cur.execute("INSERT INTO owners (name) VALUES (%s) ON CONFLICT DO NOTHING", (owner,))
                cur.close()
            else:
                conn.execute("""
                    INSERT OR REPLACE INTO tasks (id, content, type, date, owner, status, approval, comment, link, priority, feature_id, source, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (task_id, content, task_type, date_str, owner, status, approval, comment, link, '', feature_id, 'Marketing Calendar', now, now))
                conn.execute("DELETE FROM task_platforms WHERE task_id = ?", (task_id,))
                for p in platforms:
                    conn.execute("INSERT OR IGNORE INTO task_platforms (task_id, platform) VALUES (?, ?)", (task_id, p))
                    conn.execute("INSERT OR IGNORE INTO platforms (name) VALUES (?)", (p,))
                if owner:
                    conn.execute("INSERT OR IGNORE INTO owners (name) VALUES (?)", (owner,))

            imported_count += 1

    if DATABASE_URL:
        cur = conn.cursor()
        cur.execute("INSERT INTO import_log (filename, imported_at, row_count) VALUES (%s, %s, %s)",
                    (os.path.basename(filepath), now, imported_count))
        cur.close()
    else:
        conn.execute("INSERT INTO import_log (filename, imported_at, row_count) VALUES (?, ?, ?)",
                     (os.path.basename(filepath), now, imported_count))
    conn.commit()
    conn.close()
    return imported_count


def is_imported():
    conn = get_connection()
    count = _query_scalar(conn, "SELECT COUNT(*) FROM import_log", "SELECT COUNT(*) FROM import_log")
    conn.close()
    return count > 0


# --- CRUD Operations ---

def get_all_tasks():
    conn = get_connection()
    tasks = _query_all(conn, "SELECT * FROM tasks ORDER BY date ASC", "SELECT * FROM tasks ORDER BY date ASC")
    for t in tasks:
        plats = _query_all(conn,
            "SELECT platform FROM task_platforms WHERE task_id = %s",
            "SELECT platform FROM task_platforms WHERE task_id = ?", (t['id'],))
        t['platforms'] = [p['platform'] for p in plats]
    conn.close()
    return tasks


def get_task(task_id):
    conn = get_connection()
    t = _query_one(conn, "SELECT * FROM tasks WHERE id = %s", "SELECT * FROM tasks WHERE id = ?", (task_id,))
    if not t:
        conn.close()
        return None
    plats = _query_all(conn,
        "SELECT platform FROM task_platforms WHERE task_id = %s",
        "SELECT platform FROM task_platforms WHERE task_id = ?", (task_id,))
    t['platforms'] = [p['platform'] for p in plats]
    conn.close()
    return t


def create_task(task_data):
    conn = get_connection()
    now = datetime.now().isoformat()
    params = (task_data['id'], task_data['content'], task_data.get('type', ''),
              task_data.get('date', ''), task_data.get('owner', ''),
              task_data.get('status', 'To Do'), task_data.get('approval', 'Draft'),
              task_data.get('comment', ''), task_data.get('link', ''),
              task_data.get('priority', ''), task_data.get('feature_id', ''), 'Manual', now, now)
    if DATABASE_URL:
        cur = conn.cursor()
        cur.execute("""INSERT INTO tasks (id, content, type, date, owner, status, approval, comment, link, priority, feature_id, source, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", params)
        for p in task_data.get('platforms', []):
            cur.execute("INSERT INTO task_platforms (task_id, platform) VALUES (%s, %s) ON CONFLICT DO NOTHING", (task_data['id'], p))
            cur.execute("INSERT INTO platforms (name) VALUES (%s) ON CONFLICT DO NOTHING", (p,))
        if task_data.get('owner'):
            cur.execute("INSERT INTO owners (name) VALUES (%s) ON CONFLICT DO NOTHING", (task_data['owner'],))
        cur.close()
    else:
        conn.execute("""INSERT INTO tasks (id, content, type, date, owner, status, approval, comment, link, priority, feature_id, source, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""", params)
        for p in task_data.get('platforms', []):
            conn.execute("INSERT OR IGNORE INTO task_platforms (task_id, platform) VALUES (?, ?)", (task_data['id'], p))
            conn.execute("INSERT OR IGNORE INTO platforms (name) VALUES (?)", (p,))
        if task_data.get('owner'):
            conn.execute("INSERT OR IGNORE INTO owners (name) VALUES (?)", (task_data['owner'],))
    conn.commit()
    conn.close()


def update_task(task_id, task_data):
    conn = get_connection()
    now = datetime.now().isoformat()
    params = (task_data['content'], task_data.get('type', ''), task_data.get('date', ''),
              task_data.get('owner', ''), task_data.get('status', 'To Do'),
              task_data.get('approval', 'Draft'), task_data.get('comment', ''),
              task_data.get('link', ''), task_data.get('priority', ''),
              task_data.get('feature_id', ''), now, task_id)
    if DATABASE_URL:
        cur = conn.cursor()
        cur.execute("""UPDATE tasks SET content=%s, type=%s, date=%s, owner=%s, status=%s, approval=%s, comment=%s, link=%s, priority=%s, feature_id=%s, updated_at=%s WHERE id=%s""", params)
        cur.execute("DELETE FROM task_platforms WHERE task_id = %s", (task_id,))
        for p in task_data.get('platforms', []):
            cur.execute("INSERT INTO task_platforms (task_id, platform) VALUES (%s, %s) ON CONFLICT DO NOTHING", (task_id, p))
            cur.execute("INSERT INTO platforms (name) VALUES (%s) ON CONFLICT DO NOTHING", (p,))
        if task_data.get('owner'):
            cur.execute("INSERT INTO owners (name) VALUES (%s) ON CONFLICT DO NOTHING", (task_data['owner'],))
        cur.close()
    else:
        conn.execute("""UPDATE tasks SET content=?, type=?, date=?, owner=?, status=?, approval=?, comment=?, link=?, priority=?, feature_id=?, updated_at=? WHERE id=?""", params)
        conn.execute("DELETE FROM task_platforms WHERE task_id = ?", (task_id,))
        for p in task_data.get('platforms', []):
            conn.execute("INSERT OR IGNORE INTO task_platforms (task_id, platform) VALUES (?, ?)", (task_id, p))
            conn.execute("INSERT OR IGNORE INTO platforms (name) VALUES (?)", (p,))
        if task_data.get('owner'):
            conn.execute("INSERT OR IGNORE INTO owners (name) VALUES (?)", (task_data['owner'],))
    conn.commit()
    conn.close()


def delete_task(task_id):
    conn = get_connection()
    if DATABASE_URL:
        cur = conn.cursor()
        cur.execute("DELETE FROM task_platforms WHERE task_id = %s", (task_id,))
        cur.execute("DELETE FROM tasks WHERE id = %s", (task_id,))
        cur.close()
    else:
        conn.execute("DELETE FROM task_platforms WHERE task_id = ?", (task_id,))
        conn.execute("DELETE FROM tasks WHERE id = ?", (task_id,))
    conn.commit()
    conn.close()


def update_task_status(task_id, new_status):
    conn = get_connection()
    now = datetime.now().isoformat()
    if DATABASE_URL:
        cur = conn.cursor()
        cur.execute("UPDATE tasks SET status=%s, updated_at=%s WHERE id=%s", (new_status, now, task_id))
        cur.close()
    else:
        conn.execute("UPDATE tasks SET status=?, updated_at=? WHERE id=?", (new_status, now, task_id))
    conn.commit()
    conn.close()


def get_all_platforms():
    conn = get_connection()
    rows = _query_all(conn, "SELECT name FROM platforms ORDER BY name", "SELECT name FROM platforms ORDER BY name")
    conn.close()
    return [r['name'] for r in rows]


def get_all_owners():
    conn = get_connection()
    rows = _query_all(conn, "SELECT name FROM owners ORDER BY name", "SELECT name FROM owners ORDER BY name")
    conn.close()
    return [r['name'] for r in rows]


def add_platform(name):
    conn = get_connection()
    _upsert_ignore(conn, "INSERT INTO platforms (name) VALUES (%s)", "INSERT OR IGNORE INTO platforms (name) VALUES (?)", (name,))
    conn.commit()
    conn.close()


def add_owner(name):
    conn = get_connection()
    _upsert_ignore(conn, "INSERT INTO owners (name) VALUES (%s)", "INSERT OR IGNORE INTO owners (name) VALUES (?)", (name,))
    conn.commit()
    conn.close()


def get_stats():
    conn = get_connection()
    stats = {}
    stats['total'] = _query_scalar(conn, "SELECT COUNT(*) FROM tasks", "SELECT COUNT(*) FROM tasks")
    stats['approved'] = _query_scalar(conn, "SELECT COUNT(*) FROM tasks WHERE approval='Approved'", "SELECT COUNT(*) FROM tasks WHERE approval='Approved'")
    stats['pending'] = _query_scalar(conn, "SELECT COUNT(*) FROM tasks WHERE approval IN ('To be approved','In Progress')", "SELECT COUNT(*) FROM tasks WHERE approval IN ('To be approved','In Progress')")
    stats['rejected'] = _query_scalar(conn, "SELECT COUNT(*) FROM tasks WHERE approval='Rejected'", "SELECT COUNT(*) FROM tasks WHERE approval='Rejected'")
    stats['done'] = _query_scalar(conn, "SELECT COUNT(*) FROM tasks WHERE status='Done'", "SELECT COUNT(*) FROM tasks WHERE status='Done'")
    stats['in_progress'] = _query_scalar(conn, "SELECT COUNT(*) FROM tasks WHERE status='In Progress'", "SELECT COUNT(*) FROM tasks WHERE status='In Progress'")
    stats['todo'] = _query_scalar(conn, "SELECT COUNT(*) FROM tasks WHERE status='To Do'", "SELECT COUNT(*) FROM tasks WHERE status='To Do'")
    stats['blocked'] = _query_scalar(conn, "SELECT COUNT(*) FROM tasks WHERE status='Blocked'", "SELECT COUNT(*) FROM tasks WHERE status='Blocked'")
    rows = _query_all(conn,
        "SELECT owner, COUNT(*) as c FROM tasks WHERE owner != '' GROUP BY owner ORDER BY c DESC",
        "SELECT owner, COUNT(*) as c FROM tasks WHERE owner != '' GROUP BY owner ORDER BY c DESC")
    stats['by_owner'] = [(r['owner'], r['c']) for r in rows]
    rows = _query_all(conn,
        "SELECT platform, COUNT(*) as c FROM task_platforms GROUP BY platform ORDER BY c DESC",
        "SELECT platform, COUNT(*) as c FROM task_platforms GROUP BY platform ORDER BY c DESC")
    stats['by_platform'] = [(r['platform'], r['c']) for r in rows]
    conn.close()
    return stats


# --- Feature CRUD ---

def get_all_features():
    conn = get_connection()
    rows = _query_all(conn, "SELECT * FROM features ORDER BY release_date ASC", "SELECT * FROM features ORDER BY release_date ASC")
    conn.close()
    return rows


def create_feature(feature_data):
    conn = get_connection()
    now = datetime.now().isoformat()
    if DATABASE_URL:
        cur = conn.cursor()
        cur.execute("INSERT INTO features (id, name, release_date, comments, created_at) VALUES (%s, %s, %s, %s, %s)",
                    (feature_data['id'], feature_data['name'], feature_data.get('release_date', ''), feature_data.get('comments', ''), now))
        cur.close()
    else:
        conn.execute("INSERT INTO features (id, name, release_date, comments, created_at) VALUES (?, ?, ?, ?, ?)",
                     (feature_data['id'], feature_data['name'], feature_data.get('release_date', ''), feature_data.get('comments', ''), now))
    conn.commit()
    conn.close()


def update_feature(feature_id, feature_data):
    conn = get_connection()
    if DATABASE_URL:
        cur = conn.cursor()
        cur.execute("UPDATE features SET name=%s, release_date=%s, comments=%s WHERE id=%s",
                    (feature_data['name'], feature_data.get('release_date', ''), feature_data.get('comments', ''), feature_id))
        cur.close()
    else:
        conn.execute("UPDATE features SET name=?, release_date=?, comments=? WHERE id=?",
                     (feature_data['name'], feature_data.get('release_date', ''), feature_data.get('comments', ''), feature_id))
    conn.commit()
    conn.close()


def delete_feature(feature_id):
    conn = get_connection()
    # Unlink tasks from this feature
    if DATABASE_URL:
        cur = conn.cursor()
        cur.execute("UPDATE tasks SET feature_id='' WHERE feature_id=%s", (feature_id,))
        cur.execute("DELETE FROM features WHERE id=%s", (feature_id,))
        cur.close()
    else:
        conn.execute("UPDATE tasks SET feature_id='' WHERE feature_id=?", (feature_id,))
        conn.execute("DELETE FROM features WHERE id=?", (feature_id,))
    conn.commit()
    conn.close()


def reset_db():
    conn = get_connection()
    if DATABASE_URL:
        cur = conn.cursor()
        cur.execute("DELETE FROM task_platforms")
        cur.execute("DELETE FROM tasks")
        cur.execute("DELETE FROM import_log")
        cur.close()
    else:
        conn.executescript("""
            DELETE FROM task_platforms;
            DELETE FROM tasks;
            DELETE FROM import_log;
        """)
    conn.commit()
    conn.close()
